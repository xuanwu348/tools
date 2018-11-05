#encoding:utf-8
import sys

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Please use 'pip install openpyxl' install openpyxl")
    sys.exit(1)  

filename = "./testetset.xlsx"
try:
    wb = load_workbook(filename, read_only=False)
except FileNotFoundError:
    wb = Workbook(filename)

if not wb.worksheets:
    sheetname = input("please input sheets name")
    ws = wb[str(sheetname)]
else:
    sheetnames = list(map(lambda x:x.title, wb.worksheets))
    sheetnames = list(enumerate(sheetnames))
    sheetname = input("please choose %s[0]:" % sheetnames)
    if sheetname.isdigit():
        if int(sheetname) < len(sheetnames):
            ws = wb.worksheets[int(sheetname)]
        else:
            print("Index out")
    else:
        print("input gidigtal")
        
    

